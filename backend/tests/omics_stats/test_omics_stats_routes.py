from __future__ import annotations

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app

client = TestClient(app)


def test_omics_stats_analyze_upload_returns_results() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "01_实验组说明"
    sheet.append(["组名", "说明"])
    sheet.append(["control", "对照组"])
    sheet.append(["treated", "处理组"])
    sheet = workbook.create_sheet("AA")
    sheet.append(["feature", "control_1", "control_2", "treated_1", "treated_2"])
    sheet.append(["A", 1, 2, 4, 5])
    sheet.append(["B", 3, 3, 3, 3])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = client.post(
        "/api/v1/omics-stats/analyze",
        files={
            "file": (
                "omics.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "request": (
                '{"comparisons":[{"reference_group":"control","test_group":"treated"}],'
                '"top_n":1,"p_adjust_method":"none"}'
            )
        },
    )

    payload = response.json()
    assert response.status_code == 200
    assert payload["results"][0]["omics_type"] == "AA"
    assert payload["results"][0]["comparison"] == "treated_vs_control"
    assert payload["results"][0]["top"][0]["feature"] in {"A", "B"}


def test_omics_stats_preview_upload_returns_groups_and_comparisons() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "01_实验组说明"
    sheet.append(["组名", "说明"])
    sheet.append(["control", "对照组"])
    sheet.append(["treated", "处理组"])
    sheet = workbook.create_sheet("AA")
    sheet.append(["feature", "control_1", "control_2", "treated_1", "treated_2"])
    sheet.append(["A", 1, 2, 4, 5])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = client.post(
        "/api/v1/omics-stats/preview",
        files={
            "file": (
                "omics.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    payload = response.json()
    assert response.status_code == 200
    assert payload["can_analyze"] is True
    assert payload["groups"] == ["control", "treated"]
    assert payload["comparisons"] == [
        {"reference_group": "control", "test_group": "treated", "label": "treated_vs_control"}
    ]
    assert payload["sheets"][0]["name"] == "AA"
